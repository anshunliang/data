# -*- coding:utf-8 -*-
import sys
import xlwt
import openpyxl
from openpyxl import Workbook
import matplotlib
# 使用 matplotlib中的FigureCanvas (在使用 Qt5 Backends中 FigureCanvas继承自QtWidgets.QWidget)
from matplotlib.backends.backend_qt5agg import FigureCanvasQTAgg as FigureCanvas
import numpy as np

from PyQt5.QtWidgets import *
from PyQt5.QtCore import *
from matplotlib import pyplot as plt
from scipy.interpolate import make_interp_spline
from PyQt5 import QtCore, QtWidgets,QtGui
from PyQt5.QtWidgets import *
from PyQt5.QtCore import *
from PyQt5.QtGui import QIcon
from influxdb import InfluxDBClient
import pandas as pd
from influxdb import InfluxDBClient
#workbook = xlwt.Workbook(encoding='utf-8')
global ppp
global lp
global nn
global xp
global hs
hs=0
xp=[]
nn=""
lp=""
ppp=[]
#class Example(QWidget):



class Mythread(QThread):
    # 定义信号,定义参数为str类型
    breakSignal = pyqtSignal(int)
 
    def __init__(self, parent=None):
        super().__init__(parent)
        # 下面的初始化方法都可以，有的python版本不支持
        #  super(Mythread, self).__init__()
 
    def run(self):
            #要定义的行为，比如开始一个活动什么的
            print("start xc")
            a=0
            global xp
            for j in range(len(xp)):
                global nn
                data = openpyxl.load_workbook("%s.xlsx" % nn)
                # 取第一张表
                sheetnames = data.get_sheet_names()
                table = data.get_sheet_by_name(sheetnames[0])
                table = data.active
                nrows = table.max_row # 获得行数
                ncolumns = table.max_column # 获得行数
            
                aa=1
                for i in xp[j]:
                    table.cell(nrows+aa,1).value = i[0]
                    table.cell(nrows+aa,2).value = i[1]
                    aa+=1
                aa=1
                
                data.save("%s.xlsx" % nn)
                global lp
                lp=str(float(j/len(xp)))
                self.breakSignal.emit("进度%s" % j)
            
            lp="写入完成"
            self.breakSignal.emit("进度%s" % j)
            #replyp = QMessageBox.question(self, '提示', '数据保存完成', QMessageBox.Yes | QMessageBox.No, QMessageBox.No)
class Example(QMainWindow):
    

    def __init__(self):
        super().__init__()
        self.initUI()

    def initUI(self):
        self.resize(1000, 500)
        self.setWindowTitle('数据后处理')

        exitAction = QAction(QIcon('exit.png'), '&退出', self)       
        exitAction.setShortcut('Ctrl+Q')
        exitAction.setStatusTip('退出软件')
        exitAction.triggered.connect(self.close)

        bxitAction = QAction(QIcon('exit.png'), '&帮助', self)       
        bxitAction.setShortcut('Ctrl+H')
        bxitAction.setStatusTip('帮助文档')  #状态栏提示
        bxitAction.triggered.connect(self.show_child)

        self.statusBar().showMessage('状态栏',0)
        self.statusBar().setStyleSheet("background-color:gray")
 
        #self.statusBar()
        #self.statusBar.setStyleSheet("background-color:gray")
        menubar = self.menuBar()
        aMenu = menubar.addMenu('&菜单')
        aMenu.addAction(exitAction)
        aMenu.addAction(bxitAction)

        bMenu = menubar.addMenu('&教程')
        bMenu.addAction(bxitAction)


        #数据库地址及其端口
        # 实例化QLabel对象，文本显示
        self.ip = QLabel(self)
        # 设置文本标签的位置和大小
        self.ip.setGeometry(0, 30, 160, 20)
        str='数据库服务器地址'
        self.ip.setText(str)
        

        #ip地址输入框
        # 实例化QLineEdit对象，文本输入框,输入PLC的IP地址
        self.ip1 = QLineEdit(self)
        #self.ip1.setPlaceholderText("请输入要查询的内容")
        self.ip1.setText("192.168.1.104")
        self.ip1.setGeometry(160, 30, 130, 20)  #位置，大小


        self.dk= QLabel(self)
        self.dk.setGeometry(350, 30, 200, 20)
        str='端口'
        self.dk.setText(str)

        # 端口输入框
        self.dk1 = QLineEdit(self)
        self.dk1.setText("8086")
        self.dk1.setGeometry(400, 30, 80, 20)  #位置，大小
        
        #用户名
        self.name = QLabel(self)
        self.name.setGeometry(0, 70, 160, 20)
        str='用户名'
        self.name.setText(str)

        #用户名输入框
        self.name1 = QLineEdit(self)
        self.name1.setText("root")
        self.name1.setGeometry(160, 70, 130, 20)  #位置，大小

        #密码
        self.mima= QLabel(self)
        self.mima.setGeometry(350, 70, 200, 20)
        str='密码'
        self.mima.setText(str)

        # 密码输入框
        self.mima1 = QLineEdit(self)
        self.mima1.setText("123456")
        self.mima1.setGeometry(400, 70, 80, 20)  #位置，大小


   




        #获取数据库按钮
        self.hq=QPushButton(self)
        self.hq.setText("查询数据库名称")
        self.hq.setGeometry(530, 70,130,25)
        self.hq.clicked.connect(self.hqsjk)



        


        #下拉框选择数据类型
        #实例化QComBox对象
        #库名
        self.t1 = QComboBox(self)
        self.t1.setGeometry(700,75,200,20)
        self.t1.currentIndexChanged.connect(self.hello)   # 当选择内容改变时触发事件
        #表名
        self.t2 = QComboBox(self)
        self.t2.setGeometry(900,75,100,20)


        #sql语句标签
        self.dbnumber=QLabel(self)
        self.dbnumber.setGeometry(5,155,200,20)
        self.dbnumber.setText('sql语句')
        #sql语句输入框
        self.t99=QTextEdit(self)
        self.t99.setGeometry(150, 150, 830, 30)  #位置，大小
        self.t99.setText("select unit from ta01;")


        #输入带查找工位号
        self.gwh=QLabel(self)
        self.gwh.setGeometry(5,120,200,20)
        self.gwh.setText('工位号')
        #工位号输入框
        self.gwh1=QLineEdit(self)
        self.gwh1.setGeometry(50, 120, 60, 20)

        #起始时间输入
        self.readlen=QLabel(self)
        self.readlen.setGeometry(150,120,200,20)
        self.readlen.setText("起始时间：")
        #起始时间输入输入框
        self.time0=QLineEdit(self)
        #self.time0.setPlaceholderText("请输入要查询的内容")
        self.time0.setGeometry(220,120,250,20)
        self.time0.setText("'2020-11-12T12:47:56.439917Z'")


        #结束时间输入
        self.readlen=QLabel(self)
        self.readlen.setGeometry(500,120,200,20)
        self.readlen.setText("结束时间：")
        #结束时间输入输入框
        self.time1=QLineEdit(self)
        #self.time1.setPlaceholderText("请输入要查询的内容")
        self.time1.setGeometry(570,120,250,20)
        self.time1.setText("'2020-11-12T12:47:56.439917Z'")

        #生成sql语句
        self.start = QPushButton(self)
        self.start.setText("生成sql语句")         #按钮文本
        #self.start.move(850,120)                   #按钮位s置
        self.start.setGeometry(850,120,130,25)
        self.start.clicked.connect(self.sc)


        
        #开始读取按钮
        self.start = QPushButton(self)
        self.start.setText("读取并保存")         #按钮文本
        self.start.move(80,250)                   #按钮位s置
        self.start.clicked.connect(self.dq)



        #设置采样间隔
        self.jg=QLabel(self)
        self.jg.setGeometry(280,255,200,20)
        self.jg.setText("采样间隔:")
        #结束时间输入输入框
        self.jg1=QLineEdit(self)
        #self.time1.setPlaceholderText("请输入要查询的内容")
        self.jg1.setGeometry(380,255,100,20)
        self.jg1.setText("4")
        #计算数据行数
        self.jg2 = QPushButton(self)
        self.jg2.setText("确定")         #按钮文本
        self.jg2.setGeometry(500,255,80,25)
        self.jg2.clicked.connect(self.jss)


        #开始读取按钮
        self.start = QPushButton(self)
        self.start.setText("曲线")         #按钮文本
        self.start.move(900,250)                   #按钮位s置
        self.start.clicked.connect(self.qx)

        #设置采样间隔
        self.jd=QLabel(self)
        self.jd.setGeometry(80,355,200,20)
        self.jd.setText("temp")
 

        '''
        #用于展示读取到的数据
        self.test = QLabel(self)
        self.test.setGeometry(0, 300, 800, 200)
        self.test.setStyleSheet("background-color:gray")
        '''


        self.dbnumber1=QLabel(self)
        self.dbnumber1.setGeometry(0,550,900,20)
        self.dbnumber1.setText("select * from results where time>'2020-11-12T12:47:56.439917Z' and time<'2020-11-12T13:30:04.072731Z'; ")
        #select user_id from ad where user_id>30;
        self.show()
    
    #计算数据行数
    def jss(self):
        zxc=self.jg1.text()
        zxcc=int(zxc)
        global hs
        #str(int(hs/zxcc))
        self.jd.setText(str(int(hs/zxcc)))
    
    
    #测试函数
    def seeet(self):
        sex = self.plc_ip.text()
        self.test.setText(sex)
    
    #生成sql语句函数
    def sc(self):
        #ip2=self.ip1.toPlainText()
        #ip2=self.ip1.toPlainText()
        
        ip2=self.ip1.text()
        dk2=self.dk1.text()
        biao=self.t2.currentText()
        gwh2=self.gwh1.text()
        timea=self.time0.text()
        timeb=self.time1.text()
        s="select "+gwh2+" from "+biao+" "+"where  time>"+timea+" and time <"+timeb+";"
        self.t99.setText("select unit from ta01;")
 
    #查询数据库名称
    def hqsjk(self):
        #连接指定数据库
        #client = InfluxDBClient('192.168.1.104', 8086, 'root', '123456', 't')
        
        client = InfluxDBClient(self.ip1.text(), self.dk1.text(),self.name1.text(),self.mima1.text(),timeout =1,retries=2) # 初始化
        try:
            x=client.get_list_database()
            for i in x:
                self.t1.addItem(i['name'])
            
            client.close()
            
        except:
            reply = QMessageBox.question(self, '错误提示', '地址错误', QMessageBox.Yes | QMessageBox.No, QMessageBox.No)
        
        
        

    #获取数据库中的表
    def hello(self):
        self.t2.clear()
        #获取下拉框当前值
        #print(self.t1.currentText() )
        try:
            #client = InfluxDBClient('192.168.1.104', 8086, 'root', '123456', self.t1.currentText())
            client = InfluxDBClient(self.ip1.text(), self.dk1.text(),self.name1.text(),self.mima1.text(),self.t1.currentText()) # 初始化
            p=client.query('show measurements;', database=self.t1.currentText()).get_points()
        
        
            for i in p:
                self.t2.addItem(i['name'])
                #print(i)
                
            client.close()
        except:
            print("地址错误")

    #读取数据的函数
    def dq(self):
        #连接指定数据库
        #client = InfluxDBClient('192.168.1.104', 8086, 'root', '123456', 't') 

        #连接下拉框选择的库
        client = InfluxDBClient(self.ip1.text(), self.dk1.text(),self.name1.text(),self.mima1.text(),database=self.t1.currentText(),timeout =1,retries=2) # 初始化
        #client = InfluxDBClient(self.ip1.text(), self.dk1.text(),self.name1.text(),self.mima1.text(),database='t',timeout =1,retries=2) # 初始化
        print("kkkkk")



        #temp=pd.DataFrame(client.query("select * from results where time>'2020-11-12T12:47:56.439917Z' and time<'2020-11-12T13:30:04.072731Z';").get_points())
        #select user_id from ad where user_id>30;
        #pp=self.t99.text()

        #获取sql语句框的值
        pp=self.t99.toPlainText()
        
        temp=pd.DataFrame(client.query(pp).get_points())
        
        x=temp.values.tolist()

        workbook = xlwt.Workbook(encoding='utf-8')
        sheet1 = workbook.add_sheet('sheet1',cell_overwrite_ok=True)#给excel文件添加sheet
        sheet1.write(0,0,'时间')#在1行1列的单元格写入内容
        #sheet1.write_merge(0,3,1,1,'合并')#合并单元格
        sheet1.write(0,1,'测量值')#在1行1列的单元格写入内容

        #计算获取的数据行数
        global hs
        hs=len(x)


        global xp
        xp=list_split(x, 10000)
        #print(xp)

        wb = Workbook()
        # 激活 worksheet
        ws = wb.active
        # 数据可以直接分配到单元格中
        ws['A1'] = "实验"
        # 可以附加行，从第一列开始附加
        ws.append(["时间"," 测量值", "工位号"])
        # Python 类型会被自动转换
        #文件名
        global nn
        nn=self.gwh1.text()
        

        wb.save("%s.xlsx" % nn)
        print("long")


        
 
        # 创建线程
        thread = Mythread()
        # # 注册信号处理函数
        thread.breakSignal.connect(chuli)
        # # 启动线程
        thread.start()
        

        '''
        a=0
        for j in range(len(xp)):
            print("aaa")
            data = openpyxl.load_workbook("%s.xlsx" % nn)
            # 取第一张表
            sheetnames = data.get_sheet_names()
            table = data.get_sheet_by_name(sheetnames[0])
            table = data.active
            nrows = table.max_row # 获得行数
            ncolumns = table.max_column # 获得行数
            
            aa=1
            for i in xp[j]:
                
                table.cell(nrows+aa,1).value = i[0]
                table.cell(nrows+aa,2).value = i[1]
                aa+=1
            aa=1
            data.save("%s.xlsx" % nn)
        '''       




        '''
        global ppp
        a=1
        for i in x:
            print("pp")
            sheet1.write(a,0,i[0])#在1行1列的单元格写入内容
            sheet1.write(a,1,i[1])#在1行1列的单元格写入内容
            #sheet1.write(a,2,i[2])#在1行1列的单元格写入内容
            a+=1
            ppp.append(i[1])
            if a>10:
                break
        '''
        '''
        #文件名
        nn=self.gwh1.text()
        workbook.save('%s.xls' % nn)
        '''
        
        
        
        reply = QMessageBox.question(self, '提示', '数据正在保存，请耐性等待', QMessageBox.Yes | QMessageBox.No, QMessageBox.No)
        

    
    def qx(self):
        #曲线图
        
        global ppp
        qq=[]
        for i in range(len(ppp)):
            qq.append(i)
        #设置画布大小
        
        plt.figure(figsize=(12, 8)) 
        #x = np.array([6, 7, 8, 9, 10, 11, 12])
        x = np.array(qq)
        #y = np.array([1.53E+03, 5.92E+02, 2.04E+02, 7.24E+01, 2.72E+01, 1.10E+01, 4.70E+00]) 
        y = np.array(ppp)
        x_smooth = np.linspace(x.min(), x.max(), 3000)
        
        y_smooth = make_interp_spline(x, y)(x_smooth)
        plt.plot(x_smooth, y_smooth)
        print("jjjj")
        plt.show()
        
        #折线图
        '''
        plt.figure(figsize=(6, 6.5))
        x = np.array([6, 7, 8, 9, 10, 11, 12])
        y = np.array([1.53E+03, 5.92E+02, 2.04E+02, 7.24E+01, 2.72E+01, 1.10E+01, 4.70E+00])
        x_smooth = np.linspace(x.min(), x.max(), 300)
        y_smooth = make_interp_spline(x, y)(x_smooth)
        plt.plot(x_smooth, y_smooth)
        plt.show()
        '''

    
    #打开子窗口的函数
    def show_child(self):
        self.child_window = Child()
        self.child_window.show()
        print("打开子窗口")
        

#定义子窗口
class Child(QWidget):
    def __init__(self):
        super().__init__()
        self.setWindowTitle("使用教程")
        self.resize(800, 600)
       
       # 几个QWidgets
        self.figure = plt.figure()
        self.canvas = FigureCanvas(self.figure)
        self.button_plot = QtWidgets.QPushButton("绘制")

        # 连接事件
        self.button_plot.clicked.connect(self.ht)
        
        # 设置布局
        layout = QtWidgets.QVBoxLayout()
        layout.addWidget(self.canvas)
        layout.addWidget(self.button_plot)
        self.setLayout(layout)

        
        
    def ht(self):
        pass
        '''
        ax = self.figure.add_axes([0.1,0.1,0.8,0.8])
        ax.plot([1,2,3,4,5])
        '''
        '''
        x = np.array([6, 7, 8, 9, 10, 11, 12])
        y = np.array([1.53E+03, 5.92E+02, 2.04E+02, 7.24E+01, 2.72E+01, 1.10E+01, 4.70E+00])
        x_smooth = np.linspace(x.min(), x.max(), 300)
        y_smooth = make_interp_spline(x, y)(x_smooth)
        plt.plot(x_smooth, y_smooth)
        #plt.show()
        print("hhh")
        self.canvas.draw()
        '''
     
     
def list_split(items, n):
    return [items[i:i+n] for i in range(0, len(items), n)]


if __name__ == "__main__":
    app = QApplication(sys.argv)
    ex = Example()
    def chuli(a):
        global lp
        #ex.jd.setText(lp)
        ex.statusBar().showMessage("正在写入数据，进度：%s" % lp)

    
    sys.exit(app.exec_())